"""
rework_generator.py
====================
Generador de Flujos de Retrabajo – Proceso: Elaboración de Zapato
Ingeniería de Procesos – 2026

Uso:
    python rework_generator.py                        # usa datos de ejemplo (zapato)
    python rework_generator.py --excel input.xlsx     # lee plantilla Excel

Dependencias:
    pip install openpyxl
"""

import argparse
import sys
from pathlib import Path


# ──────────────────────────────────────────────────────────────
#  MODELO DE DATOS
# ──────────────────────────────────────────────────────────────

class ReworkFlow:
    """Un flujo de retrabajo con su razón, nombre de flujo, primer paso y paso de retorno."""
    def __init__(self, reason: str, flow_name: str, first_step: str, return_step: str):
        self.reason      = reason
        self.flow_name   = flow_name
        self.first_step  = first_step
        self.return_step = return_step

    def __repr__(self):
        return f"ReworkFlow(reason='{self.reason}', flow='{self.flow_name}')"


class MainStep:
    """Un paso del proceso principal con sus flujos de retrabajo asignados."""
    def __init__(self, name: str, position: int):
        self.name     = name
        self.position = position
        self.reworks: list[ReworkFlow] = []

    def add_rework(self, rw: ReworkFlow):
        self.reworks.append(rw)

    def generate_output(self) -> str:
        """Genera el string de retrabajo para este paso."""
        if not self.reworks:
            return ""
        parts = []
        for rw in self.reworks:
            parts.append(
                f"{self.name}-->GoToFlowPath[{rw.flow_name}/{rw.first_step}]\n"
                f"ReturnStep[{rw.return_step}]\n"
                f"Reason[{rw.reason}];"
            )
        return "\n".join(parts)


# ──────────────────────────────────────────────────────────────
#  DATOS DE EJEMPLO – PROCESO ELABORACIÓN DE ZAPATO
# ──────────────────────────────────────────────────────────────

def get_shoe_process() -> list[MainStep]:
    """
    Retorna el proceso principal de elaboración de zapato con
    sus flujos de retrabajo pre-asignados.
    """
    # Definir flujos de retrabajo disponibles
    reworks = {
        "corte": ReworkFlow(
            reason      = "Corte Defectuoso",
            flow_name   = "Retrabajar Corte",
            first_step  = "Revisar Patron",
            return_step = "Cortar Material"
        ),
        "pegado": ReworkFlow(
            reason      = "Pegado Deficiente",
            flow_name   = "Retrabajar Pegado",
            first_step  = "Despegar Suela",
            return_step = "Pegar Suela"
        ),
        "costura": ReworkFlow(
            reason      = "Costura Rota",
            flow_name   = "Retrabajar Costura",
            first_step  = "Descosturar Pieza",
            return_step = "Costura"
        ),
        "acabado": ReworkFlow(
            reason      = "Acabado Malo",
            flow_name   = "Retrabajar Acabado",
            first_step  = "Limpiar Superficie",
            return_step = "Acabado"
        ),
        "qc": ReworkFlow(
            reason      = "Fallo de Calidad",
            flow_name   = "Reproceso QC",
            first_step  = "Identificar Defecto",
            return_step = "Control de Calidad"
        ),
        "empaque": ReworkFlow(
            reason      = "Empaque Dañado",
            flow_name   = "Retrabajar Empaque",
            first_step  = "Desempacar",
            return_step = "Empaque"
        ),
    }

    # Definir pasos del proceso principal
    steps_config = [
        ("Cortar Material",      1, ["corte"]),
        ("Preparar Plantilla",   2, []),
        ("Ensamblar Upper",      3, []),
        ("Pegar Suela",          4, ["pegado"]),
        ("Costura",              5, ["costura"]),
        ("Acabado",              6, ["acabado"]),
        ("Control de Calidad",   7, ["qc"]),
        ("Empaque",              8, ["empaque"]),
    ]

    steps = []
    for name, pos, rw_keys in steps_config:
        step = MainStep(name=name, position=pos)
        for key in rw_keys:
            step.add_rework(reworks[key])
        steps.append(step)

    return steps


# ──────────────────────────────────────────────────────────────
#  LECTOR DE EXCEL
# ──────────────────────────────────────────────────────────────

def load_from_excel(path: str) -> list[MainStep]:
    """
    Lee la plantilla Rework Generator.xlsx y construye el modelo de datos.

    Espera dos hojas:
    - FlowStructures: FLOW | STEP | POSITION | REWORKS
    - FlowReworks:    REASON | REWORK FLOW | STEP REWORK | POSITION
    """
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] Instala openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Leer flujos de retrabajo
    rw_sheet = wb["FlowReworks"]
    rework_map: dict[str, ReworkFlow] = {}  # flow_name -> ReworkFlow
    current_reason = None
    current_flow   = None

    rows = list(rw_sheet.iter_rows(values_only=True))
    for row in rows[1:]:  # skip header
        reason, flow_name, step_rework, _ = row
        if reason:
            current_reason = reason
        if flow_name:
            current_flow = flow_name
        if current_flow and step_rework and current_reason and current_flow not in rework_map:
            rework_map[current_flow] = ReworkFlow(
                reason      = current_reason,
                flow_name   = current_flow,
                first_step  = step_rework,
                return_step = ""  # se rellena desde FlowStructures
            )

    # Leer proceso principal y asignar retrabajos
    fs_sheet = wb["FlowStructures"]
    all_steps: list[MainStep] = []
    seen_flows: set = set()

    fs_rows = list(fs_sheet.iter_rows(values_only=True))
    for row in fs_rows[1:]:
        flow_col, step_name, position, reworks_str = row
        if not step_name or not position:
            continue

        # Only process first occurrence of each flow (filter duplicates by flow+position)
        key = (flow_col, step_name)
        if key in seen_flows:
            continue
        seen_flows.add(key)

        step = MainStep(name=str(step_name), position=int(position))

        if reworks_str:
            # Parse each GoToFlowPath[...] block
            import re
            pattern = r"GoToFlowPath\[([^/\]]+)/([^\]]+)\]\s+ReturnStep\[([^\]]+)\]\s+Reason\[([^\]]+)\]"
            for m in re.finditer(pattern, reworks_str):
                rw_flow, rw_first, rw_return, rw_reason = m.groups()
                rw = ReworkFlow(
                    reason      = rw_reason.strip(),
                    flow_name   = rw_flow.strip(),
                    first_step  = rw_first.strip(),
                    return_step = rw_return.strip()
                )
                step.add_rework(rw)

        all_steps.append(step)

    # Sort by position
    all_steps.sort(key=lambda s: s.position)
    return all_steps


# ──────────────────────────────────────────────────────────────
#  GENERADOR DE OUTPUT
# ──────────────────────────────────────────────────────────────

def generate_output(steps: list[MainStep]) -> str:
    """Genera el string completo de retrabajos para todos los pasos."""
    blocks = []
    for step in steps:
        out = step.generate_output()
        if out:
            blocks.append(out)
    return "\n\n".join(blocks)


def print_summary(steps: list[MainStep]):
    """Imprime un resumen del proceso cargado."""
    print("\n" + "="*60)
    print("  PROCESO PRINCIPAL")
    print("="*60)
    for s in steps:
        rw_count = len(s.reworks)
        rw_info  = f"  [{rw_count} retrabajo{'s' if rw_count != 1 else ''}]" if rw_count else ""
        print(f"  {s.position:>2}. {s.name}{rw_info}")
    total_rw = sum(len(s.reworks) for s in steps)
    print(f"\n  Total pasos: {len(steps)} | Total retrabajos: {total_rw}")
    print("="*60 + "\n")


# ──────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generador de Flujos de Retrabajo – Proceso Elaboración de Zapato"
    )
    parser.add_argument(
        "--excel", "-e",
        metavar="FILE",
        help="Ruta al archivo Excel (plantilla Rework Generator.xlsx)"
    )
    parser.add_argument(
        "--output", "-o",
        metavar="FILE",
        help="Guardar output en archivo de texto"
    )
    args = parser.parse_args()

    # Cargar datos
    if args.excel:
        print(f"[INFO] Leyendo Excel: {args.excel}")
        steps = load_from_excel(args.excel)
    else:
        print("[INFO] Usando datos de ejemplo – Proceso Elaboración de Zapato")
        steps = get_shoe_process()

    # Resumen
    print_summary(steps)

    # Generar output
    result = generate_output(steps)

    print("OUTPUT GENERADO:")
    print("-" * 60)
    print(result)
    print("-" * 60)

    # Guardar si se especifica archivo
    if args.output:
        Path(args.output).write_text(result, encoding="utf-8")
        print(f"\n[OK] Output guardado en: {args.output}")
    else:
        # Guardar siempre como rework_output.txt
        out_path = Path("rework_output.txt")
        out_path.write_text(result, encoding="utf-8")
        print(f"\n[OK] Output guardado en: {out_path.resolve()}")


if __name__ == "__main__":
    main()
